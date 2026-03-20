import configparser
import csv
import subprocess
import sys
import re
import base64
import json
import time
import ssl
import socket
import queue
import threading
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from PIL import Image

CONFIG_FILE = "config.ini"

ASSETS_DIR     = Path("assets")
OUTPUT_DIR     = Path("output")
SCREENSHOT_DIR = OUTPUT_DIR / "screenshots"
THUMB_DIR      = OUTPUT_DIR / "thumbnails"
LOG_FILE       = OUTPUT_DIR / "logs" / "scan.log"
LAST_SCAN_FILE = OUTPUT_DIR / "last_scan.json"

ASSETS_DIR.mkdir(exist_ok=True)
SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
THUMB_DIR.mkdir(parents=True, exist_ok=True)
(LOG_FILE.parent).mkdir(parents=True, exist_ok=True)


def load_config() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg.read_dict({
        "general": {"domains_file": "domains.txt"},
        "author":  {"name": "Your Name", "email": ""},
        "titles":  {
            "main_title":       "Your Name \u2013 Web Inventory Report",
            "subtitle":         "Automatisierte Website-\u00dcbersicht",
            "sheet_inventory":  "Web Inventory",
            "sheet_summary":    "Zusammenfassung",
        },
        "colors":  {
            "primary":  "af001d",
            "dark":     "333333",
            "white":    "ffffff",
            "light_bg": "f5f5f5",
            "border":   "dee2e6",
        },
        "scan": {
            "workers":           "4",
            "goto_timeout":      "15000",
            "page_timeout":      "30000",
            "ssl_timeout":       "5",
            "screenshot_delay":  "1500",
            "card_width":        "240",
        },
    })
    cfg.read(CONFIG_FILE, encoding="utf-8")
    return cfg


CFG = load_config()
_c = CFG["colors"]
SRG_RED      = _c["primary"]
SRG_GRAY     = _c["dark"]
SRG_WHITE    = _c["white"]
SRG_LIGHT_BG = _c["light_bg"]
SRG_BORDER   = _c["border"]

_s = CFG["scan"]
WORKERS           = int(_s["workers"])
GOTO_TIMEOUT      = int(_s["goto_timeout"])
PAGE_TIMEOUT      = int(_s["page_timeout"])
SSL_TIMEOUT       = int(_s["ssl_timeout"])
SCREENSHOT_DELAY  = int(_s["screenshot_delay"])
CARD_WIDTH        = int(_s["card_width"])

THUMB_WIDTH  = 320
THUMB_HEIGHT = 200

# ── Security headers ──────────────────────────────────────────────────────────
SEC_HEADERS = [
    "strict-transport-security",
    "content-security-policy",
    "x-frame-options",
    "x-content-type-options",
    "x-xss-protection",
    "referrer-policy",
    "permissions-policy",
]

# ── CMS detection signatures ──────────────────────────────────────────────────
_CMS_SIGS = [
    ("WordPress",   [r"wp-content/", r"wp-includes/", r'name="generator"[^>]*WordPress', r"wp-json"]),
    ("Joomla",      [r"/media/jui/", r'name="generator"[^>]*Joomla', r"joomla_user_state"]),
    ("Drupal",      [r"/sites/default/files/", r'name="generator"[^>]*Drupal', r"Drupal\.settings"]),
    ("TYPO3",       [r"/typo3/", r'name="generator"[^>]*TYPO3', r"typo3conf"]),
    ("Shopify",     [r"cdn\.shopify\.com", r"myshopify\.com"]),
    ("Wix",         [r"wixsite\.com", r"wix-static\.net"]),
    ("Squarespace", [r"static\.squarespace\.com"]),
    ("Webflow",     [r"webflow\.io", r'data-wf-page']),
    ("Contao",      [r'name="generator"[^>]*Contao']),
    ("NEOS",        [r'name="generator"[^>]*Neos']),
    ("Magento",     [r"Mage\.Cookies", r"/skin/frontend/"]),
    ("PrestaShop",  [r"prestashop", r"/themes/.*?/css/global\.css"]),
    ("Ghost",       [r'name="generator"[^>]*Ghost', r"/ghost/"]),
    ("HubSpot",     [r"hs-scripts\.com", r"hubspot\.com/hs-fs"]),
]


def detect_cms(html: str, headers: dict) -> str:
    gen     = headers.get("x-generator", "").lower()
    powered = headers.get("x-powered-by", "").lower()
    for name, _ in _CMS_SIGS:
        if name.lower() in gen or name.lower() in powered:
            return name
    for name, patterns in _CMS_SIGS:
        for pat in patterns:
            if re.search(pat, html, re.IGNORECASE):
                return name
    return ""


# ── CDN / Hosting detection signatures ───────────────────────────────────────
_CDN_HEADER_SIGS = [
    ("Cloudflare",  {"cf-ray", "cf-cache-status"}),
    ("CloudFront",  {"x-amz-cf-id", "x-amz-cf-pop"}),
    ("Fastly",      {"x-fastly-request-id", "fastly-restarts", "x-served-by"}),
    ("Akamai",      {"x-akamai-transformed", "x-check-cacheable", "akamai-origin-hop"}),
    ("Vercel",      {"x-vercel-id", "x-vercel-cache"}),
    ("Netlify",     {"x-nf-request-id"}),
    ("BunnyCDN",    {"cdn-pullzone", "cdn-uid"}),
    ("Sucuri",      {"x-sucuri-id", "x-sucuri-cache"}),
    ("Azure CDN",   {"x-azure-ref"}),
    ("Imperva",     {"x-iinfo"}),
]
_CDN_SERVER_SIGS = [
    ("Cloudflare", "cloudflare"),
    ("Vercel",     "vercel"),
    ("Netlify",    "netlify"),
    ("KeyCDN",     "keycdn"),
    ("Akamai",     "akamaighost"),
    ("Google CDN", "gws"),
]


def detect_cdn(headers: dict) -> str:
    """Detect CDN/hosting from HTTP response headers."""
    srv = headers.get("server", "").lower()
    for name, sig in _CDN_SERVER_SIGS:
        if sig in srv:
            return name
    for name, header_set in _CDN_HEADER_SIGS:
        if any(h in headers for h in header_set):
            return name
    via = headers.get("via", "").lower()
    if "cloudfront" in via:
        return "CloudFront"
    return ""


def check_security_headers(headers: dict) -> dict:
    result = {h: headers.get(h, "") for h in SEC_HEADERS}
    result["_score"] = sum(1 for h in SEC_HEADERS if headers.get(h))
    return result


def classify_error(error_str: str, status) -> str:
    """Return a machine-readable scan_state string."""
    if not error_str:
        s = str(status) if status else ""
        return "ok" if s[:1] in ("2", "3", "4", "5") else ("error" if not s else "ok")
    e = error_str.lower()
    if any(k in e for k in ["name_not_resolved", "err_name_not_resolved",
                              "getaddrinfo failed", "nodename nor servname",
                              "no address associated", "dns_probe"]):
        return "dns_failed"
    if any(k in e for k in ["err_cert", "err_ssl", "err_tls", "certificate",
                              "sec_error_", "mozilla_pkix_error_", "tls_"]):
        return "tls_failed"
    if any(k in e for k in ["timeout", "timed out", "err_timed_out"]):
        return "timeout"
    if any(k in e for k in ["connection refused", "err_connection_refused", "econnrefused"]):
        return "connection_refused"
    if any(k in e for k in ["err_blocked", "access denied", "403 forbidden"]):
        return "blocked"
    return "error"


def log(msg):
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now()} - {msg}\n")


def sanitize(name):
    return re.sub(r'[^a-zA-Z0-9._-]', '_', name)


def load_hosts(file):
    with open(file, "r", encoding="utf-8") as f:
        return [l.strip() for l in f if l.strip() and not l.startswith("#")]


def create_thumbnail_16_10(src, dest):
    try:
        img = Image.open(src)
        w, h = img.size
        target_ratio = 16 / 10
        current_ratio = w / h
        if current_ratio > target_ratio:
            new_w = int(h * target_ratio)
            left = (w - new_w) // 2
            img = img.crop((left, 0, left + new_w, h))
        elif current_ratio < target_ratio:
            new_h = int(w / target_ratio)
            img = img.crop((0, 0, w, new_h))
        img = img.resize((THUMB_WIDTH, THUMB_HEIGHT), Image.LANCZOS)
        img.save(dest, "JPEG", quality=70, optimize=True)
    except Exception as e:
        log(f"Thumbnail error: {e}")


def image_to_base64(path):
    try:
        with open(path, "rb") as f:
            data = base64.b64encode(f.read()).decode("utf-8")
        ext = Path(path).suffix.lower()
        mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"
        return f"data:{mime};base64,{data}"
    except Exception:
        return ""


def get_cert_info(host: str) -> dict:
    """Get SSL cert info: expiry, subject_cn, san list. Tries strict then insecure."""
    result = {"expiry": "", "subject_cn": "", "san": []}

    def _parse(cert: dict):
        if not cert:
            return
        try:
            result["expiry"] = datetime.strptime(
                cert["notAfter"], "%b %d %H:%M:%S %Y %Z"
            ).strftime("%d.%m.%Y")
        except Exception:
            pass
        for rdns in cert.get("subject", ()):
            for k, v in rdns:
                if k == "commonName":
                    result["subject_cn"] = v
        for typ, val in cert.get("subjectAltName", ()):
            if typ == "DNS":
                result["san"].append(val)

    # Attempt 1: full TLS validation (returns complete cert dict for valid certs)
    try:
        ctx = ssl.create_default_context()
        with socket.create_connection((host, 443), timeout=SSL_TIMEOUT) as raw:
            with ctx.wrap_socket(raw, server_hostname=host) as ss:
                _parse(ss.getpeercertificate())
        return result
    except Exception:
        pass

    # Attempt 2: no validation — get DER bytes and parse via openssl subprocess
    try:
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with socket.create_connection((host, 443), timeout=SSL_TIMEOUT) as raw:
            with ctx.wrap_socket(raw, server_hostname=host) as ss:
                der = ss.getpeercertificate(binary_form=True)
        if der:
            pem = ssl.DER_cert_to_PEM_cert(der)
            proc = subprocess.run(
                ["openssl", "x509", "-noout", "-subject", "-dates",
                 "-ext", "subjectAltName", "-nameopt", "compat"],
                input=pem, capture_output=True, text=True, timeout=5,
            )
            for line in proc.stdout.splitlines():
                line = line.strip()
                if line.startswith("subject="):
                    m = re.search(r"CN\s*=\s*([^,/\n]+)", line)
                    if m:
                        result["subject_cn"] = m.group(1).strip()
                elif line.startswith("notAfter="):
                    try:
                        result["expiry"] = datetime.strptime(
                            line.split("=", 1)[1].strip(), "%b %d %H:%M:%S %Y %Z"
                        ).strftime("%d.%m.%Y")
                    except Exception:
                        pass
                elif "DNS:" in line:
                    result["san"].extend(re.findall(r"DNS:([^\s,]+)", line))
    except Exception:
        pass

    return result


def get_ip_info(host: str) -> tuple:
    """Return (ip_address, reverse_dns_hostname)."""
    try:
        ip = socket.gethostbyname(host)
    except Exception:
        return "", ""
    try:
        reverse = socket.gethostbyaddr(ip)[0]
    except Exception:
        reverse = ""
    return ip, reverse


def get_nameservers(host: str) -> str:
    """Return comma-separated authoritative nameservers for the domain."""
    parts = host.split(".")
    domain = ".".join(parts[-2:]) if len(parts) >= 2 else host
    try:
        result = subprocess.run(
            ["dig", "NS", domain, "+short", "+time=3", "+tries=1"],
            capture_output=True, text=True, timeout=5,
        )
        ns_list = sorted({l.strip().rstrip(".") for l in result.stdout.splitlines() if l.strip()})
        return ", ".join(ns_list[:6])
    except Exception:
        return ""


def get_asn_info(ip: str) -> dict:
    """Look up ASN / provider via Team Cymru DNS. Returns asn, asn_name, network, country."""
    result = {"asn": "", "asn_name": "", "network": "", "country": ""}
    if not ip:
        return result
    try:
        rev = ".".join(reversed(ip.split(".")))
        out = subprocess.run(
            ["dig", "TXT", f"{rev}.origin.asn.cymru.com", "+short", "+time=3", "+tries=1"],
            capture_output=True, text=True, timeout=6,
        ).stdout.strip().strip('"')
        if out:
            parts = [p.strip() for p in out.split("|")]
            if len(parts) >= 3:
                asn_num = parts[0].strip()
                result["asn"] = f"AS{asn_num}"
                result["network"] = parts[1].strip()
                result["country"] = parts[2].strip()
        # Second query: ASN name
        if result["asn"]:
            out2 = subprocess.run(
                ["dig", "TXT", f"{result['asn']}.asn.cymru.com", "+short", "+time=3", "+tries=1"],
                capture_output=True, text=True, timeout=6,
            ).stdout.strip().strip('"')
            if out2:
                parts2 = [p.strip() for p in out2.split("|")]
                if len(parts2) >= 5:
                    result["asn_name"] = parts2[-1].strip().split(",")[0].strip()
    except Exception:
        pass
    return result


def get_whois_info(host: str) -> dict:
    """Return registrar and domain expiry via python-whois. Capped at 10 s."""
    result = {"registrar": "", "domain_expiry": ""}

    def _query():
        try:
            import whois
            parts = host.split(".")
            domain = ".".join(parts[-2:]) if len(parts) >= 2 else host
            w = whois.whois(domain)
            result["registrar"] = str(w.registrar or "")[:60]
            exp = w.expiration_date
            if isinstance(exp, list):
                exp = exp[0]
            if exp and hasattr(exp, "strftime"):
                result["domain_expiry"] = exp.strftime("%d.%m.%Y")
        except Exception:
            pass

    t = threading.Thread(target=_query, daemon=True)
    t.start()
    t.join(timeout=10)
    return result


# Error keywords that indicate a cert/SSL issue worth retrying with ignore_https_errors
_CERT_ERROR_KEYWORDS = [
    "ERR_CERT_", "ERR_SSL_", "ERR_TLS_", "certificate", "ssl",
    "SEC_ERROR_", "MOZILLA_PKIX_ERROR_", "ERR_HTTP2_",
    "ERR_CONTENT_DECODING_", "ERR_EMPTY_RESPONSE",
]


def _is_ignorable_error(msg: str) -> bool:
    m = msg.lower()
    return any(k.lower() in m for k in _CERT_ERROR_KEYWORDS)


def _do_scan(page, url, host):
    """Perform goto + screenshot + security headers + redirect chain + CMS."""
    redirect_chain = []

    def on_response(resp):
        if resp.status in (301, 302, 303, 307, 308):
            redirect_chain.append({
                "url":      resp.url,
                "status":   resp.status,
                "location": resp.headers.get("location", ""),
            })

    page.on("response", on_response)
    try:
        start = time.time()
        response = page.goto(url, timeout=GOTO_TIMEOUT, wait_until="domcontentloaded")
        load_time = round(time.time() - start, 2)
        final_url = page.url
        title     = page.title()
        filename  = sanitize(host)

        screenshot_path = SCREENSHOT_DIR / f"{filename}.png"
        thumb_path      = THUMB_DIR      / f"{filename}.jpg"
        if SCREENSHOT_DELAY > 0:
            page.wait_for_timeout(SCREENSHOT_DELAY)
        # Hide common cookie/GDPR consent overlays before screenshot
        try:
            page.evaluate("""(() => {
                const s = document.createElement('style');
                s.textContent = [
                    '#onetrust-banner-sdk','#onetrust-consent-sdk',
                    '#CybotCookiebotDialog','#CookiebotWidget',
                    '.cc-window','.cc-banner','.cc-float',
                    '#cookie-notice','.cookie-notice',
                    '#cookie-banner','.cookie-banner',
                    '#cookieConsent','.cookieConsent',
                    '#cookie-law-info-bar','.cookies-eu-banner',
                    '#gdpr-banner','.gdpr-banner',
                    '.cookiebanner','.cookie-warning',
                    '[id^="cookie-consent"]','[class^="cookie-consent"]',
                    '.pea_cook_wrapper','#tarteaucitron',
                ].join(',') + '{display:none!important}' +
                'body{overflow:auto!important}';
                document.head.appendChild(s);
            })()""")
            page.wait_for_timeout(150)
        except Exception:
            pass
        page.screenshot(path=str(screenshot_path), full_page=False)
        create_thumbnail_16_10(screenshot_path, thumb_path)

        resp_headers = {}
        if response:
            resp_headers = {k.lower(): v for k, v in response.headers.items()}

        sec_headers = check_security_headers(resp_headers)
        cdn = detect_cdn(resp_headers)

        try:
            html_content = page.content()
        except Exception:
            html_content = ""
        cms = detect_cms(html_content, resp_headers)

        # ── Cert info via Playwright security_details (reuses existing TLS conn) ──
        cert_expiry = ""
        cert_subject_cn = ""
        try:
            sd = response.security_details() if response else None
            if sd:
                vt = sd.get("valid_to")
                if vt:
                    from datetime import timezone
                    cert_expiry = datetime.fromtimestamp(
                        vt, tz=timezone.utc
                    ).strftime("%d.%m.%Y")
                cert_subject_cn = sd.get("subject_name", "")
        except Exception:
            pass

        return {
            "status":           response.status if response else "",
            "final_url":        final_url,
            "title":            title,
            "load_time":        load_time,
            "screenshot":       str(screenshot_path),
            "thumbnail":        str(thumb_path),
            "security_headers": sec_headers,
            "redirect_chain":   redirect_chain,
            "cms":              cms,
            "cdn":              cdn,
            "ssl_expiry":       cert_expiry,
            "cert_subject_cn":  cert_subject_cn,
        }
    finally:
        page.remove_listener("response", on_response)


def scan_host(page, browser, host):
    scan_time = datetime.now().isoformat()

    # Sequential DNS lookups first (needed for ASN lookup)
    ip, reverse_dns = get_ip_info(host)
    nameservers = get_nameservers(host)

    # Parallel background lookups while browser scan runs
    # get_cert_info → SAN only (expiry+subject come from Playwright security_details)
    san_result   = [{"san": []}]
    asn_result   = [{"asn": "", "asn_name": "", "network": "", "country": ""}]
    whois_result = [{"registrar": "", "domain_expiry": ""}]

    def _san():
        c = get_cert_info(host)
        san_result[0]["san"] = c["san"]
    def _asn():   asn_result[0]   = get_asn_info(ip)
    def _whois(): whois_result[0] = get_whois_info(host)

    bg = [threading.Thread(target=f, daemon=True) for f in (_san, _asn, _whois)]
    for t in bg:
        t.start()

    # Browser scan
    urls = [f"https://{host}", f"http://{host}"]
    last_error = ""
    scan_result = None

    for url in urls:
        # ── Normal attempt (strict SSL) ───────────────────────────────────
        try:
            info = _do_scan(page, url, host)
            scan_result = {"host": host, **info, "error": ""}
            break
        except Exception as e:
            err_str = str(e)
            last_error = err_str
            log(f"Error on {url}: {err_str}")

            # ── Retry with ignore_https_errors for cert/SSL issues ────────
            if _is_ignorable_error(err_str):
                try:
                    ctx = browser.new_context(
                        viewport={"width": 1280, "height": 800},
                        ignore_https_errors=True,
                    )
                    p2 = ctx.new_page()
                    try:
                        info = _do_scan(p2, url, host)
                        scan_result = {"host": host, **info, "error": err_str}
                        break
                    finally:
                        ctx.close()
                except Exception as e2:
                    log(f"Retry failed for {host}: {e2}")
                    last_error = err_str  # keep original error

    # Wait for background lookups (capped)
    for t in bg:
        t.join(timeout=15)

    asn   = asn_result[0]
    whois = whois_result[0]

    # ssl_expiry + cert_subject_cn come from Playwright (_do_scan); SAN from background
    browser_ssl_expiry    = (scan_result or {}).get("ssl_expiry", "")
    browser_cert_subject  = (scan_result or {}).get("cert_subject_cn", "")
    cert_san_str          = ", ".join(san_result[0]["san"][:10])

    extra = {
        "ip":             ip,
        "reverse_dns":    reverse_dns,
        "nameservers":    nameservers,
        "ssl_expiry":     browser_ssl_expiry,
        "cert_subject_cn": browser_cert_subject,
        "cert_san":       cert_san_str,
        "asn":            asn["asn"],
        "asn_name":       asn["asn_name"],
        "hosting_country": asn["country"],
        "registrar":      whois["registrar"],
        "domain_expiry":  whois["domain_expiry"],
        "scan_time":      scan_time,
    }

    if scan_result:
        scan_result.update(extra)
        scan_result["scan_state"] = classify_error(
            scan_result.get("error", ""), scan_result.get("status")
        )
        return scan_result

    return {
        "host": host, "status": "", "final_url": "", "title": "",
        "load_time": "", "screenshot": "", "thumbnail": "",
        "security_headers": {"_score": 0}, "redirect_chain": [],
        "cms": "", "cdn": "",
        "error": last_error,
        "scan_state": classify_error(last_error, ""),
        **extra,
    }


# ── Delta Report ─────────────────────────────────────────────────────────────

def load_previous_results() -> dict:
    if not LAST_SCAN_FILE.exists():
        return {}
    try:
        with open(LAST_SCAN_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {r["host"]: r for r in data if r}
    except Exception:
        return {}


def save_results_json(data: list):
    with open(LAST_SCAN_FILE, "w", encoding="utf-8") as f:
        json.dump([r for r in data if r], f, ensure_ascii=False, default=str)


def compute_delta(result: dict, previous: dict) -> str:
    host = result["host"]
    if host not in previous:
        return "new"
    prev = previous[host]
    if (str(result.get("status", "")) != str(prev.get("status", "")) or
            result.get("title", "")     != prev.get("title", "")     or
            result.get("final_url", "") != prev.get("final_url", "")):
        return "changed"
    return "unchanged"


# ── CSV ──────────────────────────────────────────────────────────────────────

def _redirect_chain_str(chain: list) -> str:
    if not chain:
        return ""
    return " → ".join(f'{r["url"]} ({r["status"]})' for r in chain)


def _sec_score(row: dict) -> int:
    return row.get("security_headers", {}).get("_score", 0)


def write_csv(data):
    fieldnames = [
        "host", "ip", "reverse_dns", "nameservers",
        "asn", "asn_name", "hosting_country",
        "scan_state", "status", "final_url", "title", "load_time",
        "ssl_expiry", "cert_subject_cn", "cert_san",
        "cms", "cdn", "sec_score",
        "redirects", "redirect_chain",
        "registrar", "domain_expiry",
        "delta", "error", "scan_time",
    ]
    rows = []
    for r in data:
        rows.append({
            "host":            r["host"],
            "ip":              r.get("ip", ""),
            "reverse_dns":     r.get("reverse_dns", ""),
            "nameservers":     r.get("nameservers", ""),
            "asn":             r.get("asn", ""),
            "asn_name":        r.get("asn_name", ""),
            "hosting_country": r.get("hosting_country", ""),
            "scan_state":      r.get("scan_state", ""),
            "status":          r.get("status", ""),
            "final_url":       r.get("final_url", ""),
            "title":           r.get("title", ""),
            "load_time":       r.get("load_time", ""),
            "ssl_expiry":      r.get("ssl_expiry", ""),
            "cert_subject_cn": r.get("cert_subject_cn", ""),
            "cert_san":        r.get("cert_san", ""),
            "cms":             r.get("cms", ""),
            "cdn":             r.get("cdn", ""),
            "sec_score":       _sec_score(r),
            "redirects":       len(r.get("redirect_chain", [])),
            "redirect_chain":  _redirect_chain_str(r.get("redirect_chain", [])),
            "registrar":       r.get("registrar", ""),
            "domain_expiry":   r.get("domain_expiry", ""),
            "delta":           r.get("delta", ""),
            "error":           r.get("error", ""),
            "scan_time":       r.get("scan_time", ""),
        })
    with open(OUTPUT_DIR / "report.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    log("CSV report written.")


# ── Excel ────────────────────────────────────────────────────────────────────

def write_excel(data):
    _a = CFG["author"]
    _t = CFG["titles"]
    wb = Workbook()
    wb.properties.creator = _a["name"]
    wb.properties.title = _t["main_title"]
    wb.properties.subject = _t["subtitle"]
    wb.properties.description = (
        f"Erstellt mit Web Inventory Reporter\n"
        f"{_a['name']}\n"
        f"{_a['email']}"
    )
    wb.properties.company = ""

    # ── Styles ──
    hdr_font = Font(name="Arial", bold=True, color=SRG_WHITE, size=11)
    hdr_fill = PatternFill("solid", fgColor=SRG_RED)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color=SRG_BORDER),
        right=Side(style="thin", color=SRG_BORDER),
        top=Side(style="thin", color=SRG_BORDER),
        bottom=Side(style="thin", color=SRG_BORDER),
    )
    zebra = PatternFill("solid", fgColor=SRG_LIGHT_BG)
    white = PatternFill("solid", fgColor=SRG_WHITE)
    font_ok = Font(name="Arial", size=10, bold=True, color="198754")
    font_redir = Font(name="Arial", size=10, bold=True, color="856404")
    font_err = Font(name="Arial", size=10, bold=True, color=SRG_RED)
    title_font = Font(name="Arial", size=14, bold=True, color=SRG_RED)
    subtitle_font = Font(name="Arial", size=10, color=SRG_GRAY)
    label_font = Font(name="Arial", size=10, bold=True, color=SRG_GRAY)
    value_font = Font(name="Arial", size=11, color=SRG_GRAY)

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 1: Zusammenfassung
    # ═══════════════════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = _t["sheet_summary"]
    ws_sum.sheet_properties.tabColor = SRG_RED

    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    total = len(data)
    ok = sum(1 for d in data if str(d.get("status", "")).startswith("2"))
    redirects = sum(1 for d in data if str(d.get("status", "")).startswith("3"))
    client_err = sum(1 for d in data if str(d.get("status", "")).startswith("4"))
    server_err = sum(1 for d in data if str(d.get("status", "")).startswith("5"))
    scan_err = sum(1 for d in data if d.get("error"))
    load_times = [float(d["load_time"]) for d in data if d.get("load_time")]
    avg_load = sum(load_times) / len(load_times) if load_times else 0
    max_load = max(load_times) if load_times else 0
    min_load = min(load_times) if load_times else 0

    ws_sum.column_dimensions["A"].width = 5
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 18
    ws_sum.column_dimensions["D"].width = 28
    ws_sum.column_dimensions["E"].width = 18

    ws_sum["B2"] = _t["main_title"]
    ws_sum["B2"].font = title_font
    ws_sum["B3"] = f"Erstellt: {ts}"
    ws_sum["B3"].font = subtitle_font
    ws_sum["B4"] = f"{_a['name']} · {_a['email']}"
    ws_sum["B4"].font = Font(name="Arial", size=9, color="6c757d")

    r = 6
    stats = [
        ("Total Hosts", total),
        ("Status 2xx (OK)", ok),
        ("Status 3xx (Redirect)", redirects),
        ("Status 4xx (Client Error)", client_err),
        ("Status 5xx (Server Error)", server_err),
        ("Scan-Fehler", scan_err),
        ("", ""),
        ("⌀ Ladezeit", f"{avg_load:.2f}s"),
        ("Min. Ladezeit", f"{min_load:.2f}s"),
        ("Max. Ladezeit", f"{max_load:.2f}s"),
    ]
    for label, val in stats:
        if label == "":
            r += 1
            continue
        ws_sum.cell(row=r, column=2, value=label).font = label_font
        ws_sum.cell(row=r, column=3, value=val).font = value_font
        ws_sum.cell(row=r, column=3).alignment = center_align
        r += 1

    # Top 10 langsamste
    r += 1
    ws_sum.cell(row=r, column=2, value="Top 10 langsamste Hosts").font = title_font
    r += 1
    for hdr_col, hdr_txt in [(2, "Host"), (3, "Ladezeit"), (4, "Status")]:
        c = ws_sum.cell(row=r, column=hdr_col, value=hdr_txt)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border
    r += 1

    slowest = sorted(
        [d for d in data if d.get("load_time")],
        key=lambda x: float(x["load_time"]), reverse=True
    )[:10]
    for d in slowest:
        ws_sum.cell(row=r, column=2, value=d["host"]).font = cell_font
        ws_sum.cell(row=r, column=2).border = border
        ws_sum.cell(row=r, column=3, value=f'{d["load_time"]}s').font = cell_font
        ws_sum.cell(row=r, column=3).alignment = center_align
        ws_sum.cell(row=r, column=3).border = border
        ws_sum.cell(row=r, column=4, value=d["status"]).font = cell_font
        ws_sum.cell(row=r, column=4).alignment = center_align
        ws_sum.cell(row=r, column=4).border = border
        r += 1

    # SSL bald ablaufend
    r += 1
    ws_sum.cell(row=r, column=2, value="SSL-Zertifikate — bald ablaufend").font = title_font
    r += 1
    for hdr_col, hdr_txt in [(2, "Host"), (3, "Ablaufdatum"), (4, "Verbleibend")]:
        c = ws_sum.cell(row=r, column=hdr_col, value=hdr_txt)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border
    r += 1

    ssl_hosts = []
    for d in data:
        if d.get("ssl_expiry"):
            try:
                exp = datetime.strptime(d["ssl_expiry"], "%d.%m.%Y")
                days = (exp - datetime.now()).days
                ssl_hosts.append((d["host"], d["ssl_expiry"], days))
            except Exception:
                pass
    ssl_hosts.sort(key=lambda x: x[2])
    for host, expiry, days in ssl_hosts[:15]:
        ws_sum.cell(row=r, column=2, value=host).font = cell_font
        ws_sum.cell(row=r, column=2).border = border
        ws_sum.cell(row=r, column=3, value=expiry).font = cell_font
        ws_sum.cell(row=r, column=3).alignment = center_align
        ws_sum.cell(row=r, column=3).border = border
        days_cell = ws_sum.cell(row=r, column=4, value=f"{days} Tage")
        days_cell.font = font_err if days < 30 else (font_redir if days < 90 else font_ok)
        days_cell.alignment = center_align
        days_cell.border = border
        r += 1

    # Fehlerhafte Hosts
    err_hosts = [d for d in data if d.get("error")]
    if err_hosts:
        r += 1
        ws_sum.cell(row=r, column=2, value="Hosts mit Scan-Fehlern").font = title_font
        r += 1
        for hdr_col, hdr_txt in [(2, "Host"), (3, "Fehler")]:
            c = ws_sum.cell(row=r, column=hdr_col, value=hdr_txt)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = border
        ws_sum.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1
        for d in err_hosts[:30]:
            ws_sum.cell(row=r, column=2, value=d["host"]).font = cell_font
            ws_sum.cell(row=r, column=2).border = border
            err_c = ws_sum.cell(row=r, column=3, value=d["error"][:80])
            err_c.font = Font(name="Arial", size=9, color=SRG_RED)
            err_c.border = border
            ws_sum.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # Sheet 2: Web Inventory (Daten)
    # ═══════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet(_t["sheet_inventory"])

    # Col: Host IP RevDNS NS ASN Provider Country ScanState Status Ladezeit SSL CertSubj CertSAN CMS CDN Sec Redir Registrar DomainExp FinalURL Titel Delta Fehler ScanTime Preview
    headers    = [
        "Preview",
        "Host", "IP-Adresse", "Reverse DNS", "Nameserver",
        "ASN", "Provider", "Land",
        "Scan State", "Status", "Ladezeit (s)", "SSL gültig bis",
        "Cert Subject", "Cert SAN",
        "CMS", "CDN", "Sec.", "Redirects",
        "Registrar", "Domain-Ablauf",
        "Final URL", "Titel", "Delta", "Fehler", "Scan-Zeit",
    ]
    col_widths = [
        28,
        28, 16, 22, 32,
        12, 24, 8,
        14, 10, 14, 16,
        24, 30,
        14, 14, 8, 12,
        24, 16,
        40, 30, 10, 30, 20,
    ]

    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
    ws.row_dimensions[1].height = 32

    for j, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"

    _delta_colors = {"new": "198754", "changed": "856404", "unchanged": SRG_GRAY}
    _preview_col  = 1  # first column = Preview (col A)

    for i, row in enumerate(data, start=2):
        fill     = zebra if i % 2 == 0 else white
        sec_val  = f'{_sec_score(row)}/7'
        redir_val = _redirect_chain_str(row.get("redirect_chain", []))[:60] or "—"
        delta_val = row.get("delta", "")

        vals = [
            "",  # col 1 = Preview (image inserted separately below)
            row["host"], row.get("ip", ""), row.get("reverse_dns", ""), row.get("nameservers", ""),
            row.get("asn", ""), row.get("asn_name", ""), row.get("hosting_country", ""),
            row.get("scan_state", ""),
            row.get("status", ""), row.get("load_time", ""), row.get("ssl_expiry", ""),
            row.get("cert_subject_cn", ""), row.get("cert_san", ""),
            row.get("cms", "") or "—", row.get("cdn", "") or "—",
            sec_val, redir_val,
            row.get("registrar", ""), row.get("domain_expiry", ""),
            row.get("final_url", ""), row.get("title", ""), delta_val,
            row.get("error", ""), row.get("scan_time", ""),
        ]
        # Column indices (1-based):
        # 1=Preview 2=Host 3=IP 4=RevDNS 5=NS 6=ASN 7=Provider 8=Country
        # 9=ScanState 10=Status 11=Ladezeit 12=SSL 13=CertSubj 14=CertSAN
        # 15=CMS 16=CDN 17=Sec 18=Redir 19=Registrar 20=DomainExp
        # 21=FinalURL 22=Titel 23=Delta 24=Fehler 25=ScanTime

        for j, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font, c.fill, c.border = cell_font, fill, border
            # center: ScanState(9), Status(10), Ladezeit(11), SSL(12), Sec(17), Redir(18), Delta(23), ScanTime(25)
            c.alignment = center_align if j in (9, 10, 11, 12, 17, 18, 23, 25) else cell_align

            if j == 10 and val:  # Status
                s = int(val) if str(val).isdigit() else 0
                c.font = font_ok if 200 <= s < 300 else (font_redir if 300 <= s < 400 else font_err)
            elif j == 17:  # Sec score
                score = _sec_score(row)
                c.font = Font(name="Arial", size=10, bold=True,
                              color="198754" if score >= 6 else ("856404" if score >= 4 else SRG_RED))
            elif j == 23 and val:  # Delta
                col = _delta_colors.get(val, SRG_GRAY)
                c.font = Font(name="Arial", size=10, bold=True, color=col)

        if row.get("thumbnail") and Path(row["thumbnail"]).exists():
            try:
                img = XLImage(row["thumbnail"])
                img.width, img.height = 160, 100
                ws.add_image(img, f"{get_column_letter(_preview_col)}{i}")
            except Exception as e:
                log(f"Excel image error for {row['host']}: {e}")

        ws.row_dimensions[i].height = 80

    # Conditional formatting: Ladezeit (now column K, index 11)
    lt_range = f"K2:K{len(data)+1}"
    ws.conditional_formatting.add(lt_range, CellIsRule(
        operator="greaterThan", formula=["3"],
        fill=PatternFill("solid", fgColor="fde8e8"),
        font=Font(name="Arial", size=10, bold=True, color=SRG_RED)
    ))
    ws.conditional_formatting.add(lt_range, CellIsRule(
        operator="between", formula=["1.5", "3"],
        fill=PatternFill("solid", fgColor="fff3cd"),
        font=Font(name="Arial", size=10, bold=True, color="856404")
    ))
    ws.conditional_formatting.add(lt_range, CellIsRule(
        operator="lessThanOrEqual", formula=["1.5"],
        fill=PatternFill("solid", fgColor="e6f9ed"),
        font=Font(name="Arial", size=10, bold=True, color="198754")
    ))

    wb.save(OUTPUT_DIR / "report.xlsx")
    log("Excel report written.")


# ── HTML ─────────────────────────────────────────────────────────────────────

def status_badge(status):
    if not status:
        return '<span class="badge badge-unknown">N/A</span>'
    s = int(status) if str(status).isdigit() else 0
    if 200 <= s < 300:   cls = "badge-ok"
    elif 300 <= s < 400: cls = "badge-redirect"
    elif 400 <= s < 500: cls = "badge-client-err"
    else:                cls = "badge-server-err"
    return f'<span class="badge {cls}">{status}</span>'


def lt_indicator(lt):
    if not lt: return "—"
    t = float(lt)
    cls = "lt-fast" if t < 1.5 else ("lt-medium" if t < 3.0 else "lt-slow")
    return f'<span class="lt {cls}">{t:.2f}s</span>'


def write_html(data):
    _a = CFG["author"]
    _t = CFG["titles"]
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    total = len(data)
    ok = sum(1 for d in data if str(d.get("status", "")).startswith("2"))
    redirects = sum(1 for d in data if str(d.get("status", "")).startswith("3"))
    errors = sum(1 for d in data if d.get("error"))
    load_times = [float(d["load_time"]) for d in data if d.get("load_time")]
    avg_load = sum(load_times) / len(load_times) if load_times else 0

    # Thumbnail for header (assets/thumbnail.png)
    thumb_asset = ASSETS_DIR / "thumbnail.png"
    thumb_b64   = image_to_base64(str(thumb_asset)) if thumb_asset.exists() else ""

    # Build JSON data array for client-side pagination/search/sort
    json_data = []
    for row in data:
        b64   = image_to_base64(row["thumbnail"]) if row.get("thumbnail") else ""
        chain = row.get("redirect_chain", [])
        sh    = row.get("security_headers", {})
        json_data.append({
            "host":      row["host"],
            "status":    str(row.get("status", "")),
            "final_url": row.get("final_url", ""),
            "title":     row.get("title", ""),
            "load_time": row["load_time"] if row.get("load_time") else None,
            "ssl_expiry": row.get("ssl_expiry", ""),
            "thumb":     b64,
            "error":     row.get("error", ""),
            "cms":       row.get("cms", ""),
            "cdn":       row.get("cdn", ""),
            "scan_state": row.get("scan_state", ""),
            "cert_subject_cn": row.get("cert_subject_cn", ""),
            "cert_san":  row.get("cert_san", ""),
            "asn":       row.get("asn", ""),
            "asn_name":  row.get("asn_name", ""),
            "hosting_country": row.get("hosting_country", ""),
            "registrar": row.get("registrar", ""),
            "domain_expiry": row.get("domain_expiry", ""),
            "scan_time": row.get("scan_time", ""),
            "sec_score": sh.get("_score", 0),
            "sec_detail": {k: bool(v) for k, v in sh.items() if k != "_score"},
            "redirect_count": len(chain),
            "redirect_chain": [f'{r["url"]} ({r["status"]})' for r in chain],
            "delta":     row.get("delta", ""),
            "ip":        row.get("ip", ""),
            "reverse_dns": row.get("reverse_dns", ""),
            "nameservers": row.get("nameservers", ""),
        })

    data_json = json.dumps(json_data, ensure_ascii=False).replace('</', '<\\/')

    html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{_t['main_title']}</title>
<style>
:root {{
    --srg-red: #af001d;
    --srg-red-dark: #8a0017;
    --srg-red-light: #fdeaed;
    --srg-gray: #333333;
    --srg-gray-light: #6c757d;
    --srg-bg: #f5f5f5;
    --srg-white: #ffffff;
    --srg-border: #dee2e6;
    --srg-green: #198754;
    --srg-green-bg: #e6f9ed;
    --srg-yellow: #856404;
    --srg-yellow-bg: #fff3cd;
    --r: 8px;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,Arial,sans-serif; background:var(--srg-bg); color:var(--srg-gray); line-height:1.5; -webkit-font-smoothing:antialiased; }}

.header {{ background:var(--srg-red); color:var(--srg-white); padding:1.6rem 2.5rem; display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:1rem; }}
.header h1 {{ font-size:1.5rem; font-weight:700; letter-spacing:-0.02em; }}
.header .sub {{ font-size:0.85rem; opacity:0.85; margin-top:0.15rem; }}
.header-meta {{ font-size:0.85rem; opacity:0.8; text-align:right; }}

.stats {{ display:flex; gap:0.6rem; padding:0.8rem 2.5rem; background:var(--srg-white); border-bottom:2px solid var(--srg-red); flex-wrap:wrap; }}
.stat {{ display:flex; align-items:center; gap:0.35rem; padding:0.3rem 0.8rem; border-radius:var(--r); font-size:0.82rem; font-weight:500; }}
.stat .n {{ font-size:1.15rem; font-weight:700; }}
.stat-total {{ background:var(--srg-red-light); color:var(--srg-red); }}
.stat-ok {{ background:var(--srg-green-bg); color:var(--srg-green); }}
.stat-redir {{ background:var(--srg-yellow-bg); color:var(--srg-yellow); }}
.stat-err {{ background:#fde8e8; color:var(--srg-red); }}
.stat-speed {{ background:#edf2f7; color:var(--srg-gray); }}

.toolbar {{ display:flex; gap:0.5rem; padding:0.8rem 2.5rem; align-items:center; flex-wrap:wrap; }}
.toolbar button {{ padding:0.4rem 1rem; border:1px solid var(--srg-border); border-radius:var(--r); background:var(--srg-white); font-size:0.82rem; cursor:pointer; transition:all 0.15s; font-family:inherit; }}
.toolbar button.active {{ background:var(--srg-red); color:var(--srg-white); border-color:var(--srg-red); }}
.toolbar button:hover:not(.active) {{ border-color:var(--srg-red); color:var(--srg-red); }}
.search-box {{ flex:1; min-width:200px; max-width:400px; padding:0.4rem 0.8rem; border:1px solid var(--srg-border); border-radius:var(--r); font-size:0.85rem; font-family:inherit; outline:none; transition:border-color 0.15s; }}
.search-box:focus {{ border-color:var(--srg-red); }}
.filter-select {{ padding:0.4rem 0.6rem; border:1px solid var(--srg-border); border-radius:var(--r); font-size:0.82rem; font-family:inherit; background:var(--srg-white); }}

.cards {{ display:grid; grid-template-columns:repeat(auto-fill,minmax({CARD_WIDTH}px,1fr)); gap:1rem; padding:0.8rem 2.5rem 1rem; }}
.card {{ background:var(--srg-white); border-radius:var(--r); overflow:visible; border:1px solid var(--srg-border); transition:box-shadow 0.2s,transform 0.15s; position:relative; }}
.card:hover {{ box-shadow:0 4px 16px rgba(0,0,0,0.07); transform:translateY(-1px); }}
.card-image {{ background:#eaeaea; aspect-ratio:16/10; overflow:hidden; display:flex; align-items:flex-start; border-radius:var(--r) var(--r) 0 0; }}
.card-image img {{ width:100%; height:100%; object-fit:cover; object-position:top; }}
.no-preview {{ display:flex; align-items:center; justify-content:center; width:100%; color:var(--srg-gray-light); font-size:0.85rem; }}
.card-body {{ padding:0.7rem 1rem 0.8rem; }}
.card-title {{ font-size:0.95rem; font-weight:700; margin-bottom:0.25rem; }}
.card-meta {{ display:flex; align-items:center; gap:0.4rem; margin-bottom:0.25rem; flex-wrap:wrap; }}
.card-pagetitle {{ font-size:0.78rem; color:var(--srg-gray-light); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; margin-bottom:0.15rem; }}
.card-url {{ font-size:0.75rem; color:var(--srg-red); text-decoration:none; word-break:break-all; display:block; margin-bottom:0.15rem; }}
.card-url:hover {{ text-decoration:underline; }}
.card-ssl {{ font-size:0.72rem; color:var(--srg-green); }}
.card-dns {{ margin-top:0.2rem; font-size:0.7rem; color:var(--srg-gray-light); line-height:1.5; }}
.card-ip {{ cursor:default; }}
.card-ns {{ }}
.card-asn {{ }}
.cell-asn {{ font-size:0.75rem; color:var(--srg-gray-light); }}
.filter-link {{ cursor:pointer; border-radius:3px; }}
.filter-link:hover {{ opacity:0.7; text-decoration:underline; outline:1px dashed currentColor; }}
.filter-active-bar {{ display:flex; align-items:center; gap:0.6rem; padding:0.45rem 2.5rem; background:var(--srg-yellow-bg); border-bottom:1px solid #ffe69c; font-size:0.82rem; color:var(--srg-yellow); }}
.filter-active-bar .f-tag {{ background:var(--srg-yellow); color:#fff; padding:0.15rem 0.55rem; border-radius:4px; font-weight:700; font-size:0.78rem; display:inline-flex; align-items:center; gap:3px; }}
.filter-active-bar .f-x {{ cursor:pointer; opacity:0.75; font-weight:400; font-size:0.9em; }}
.filter-active-bar .f-x:hover {{ opacity:1; }}
.filter-active-bar button {{ padding:0.15rem 0.6rem; border:1px solid var(--srg-yellow); border-radius:4px; background:transparent; color:var(--srg-yellow); cursor:pointer; font-size:0.78rem; font-family:inherit; }}
.filter-active-bar button:hover {{ background:var(--srg-yellow); color:#fff; }}
.card-error {{ margin-top:0.4rem; padding:0.3rem 0.5rem; background:#fde8e8; border-radius:4px; font-size:0.72rem; color:var(--srg-red); }}
.cell-ip {{ font-size:0.78rem; font-family:monospace; white-space:nowrap; }}
.cell-ns {{ font-size:0.75rem; color:var(--srg-gray-light); }}

.badge {{ display:inline-block; padding:0.1rem 0.45rem; border-radius:4px; font-size:0.72rem; font-weight:600; white-space:nowrap; }}
.badge-ok {{ background:var(--srg-green-bg); color:var(--srg-green); }}
.badge-redirect {{ background:var(--srg-yellow-bg); color:var(--srg-yellow); }}
.badge-client-err {{ background:#fde8e8; color:var(--srg-red); }}
.badge-server-err {{ background:#fde8e8; color:#8b0000; }}
.badge-unknown {{ background:#edf2f7; color:var(--srg-gray-light); }}
.badge-cms {{ background:#e8f0fe; color:#1a56db; }}
.badge-cdn {{ background:#e8f4fd; color:#0369a1; }}
.badge-new {{ background:var(--srg-green-bg); color:var(--srg-green); }}
.badge-changed {{ background:var(--srg-yellow-bg); color:var(--srg-yellow); }}
.sec-score {{ display:inline-flex; align-items:center; gap:2px; font-size:0.72rem; font-weight:700; }}
.sec-high {{ color:var(--srg-green); }}
.sec-mid {{ color:var(--srg-yellow); }}
.sec-low {{ color:var(--srg-red); }}
.card-sec {{ font-size:0.7rem; }}
.redir-chip {{ font-size:0.7rem; color:#6c757d; cursor:help; }}
.sec-tooltip {{ position:relative; cursor:help; }}
.sec-tooltip:hover .sec-tip {{ display:block; }}
.sec-tip {{ display:none; position:absolute; z-index:1000; bottom:120%; left:50%; transform:translateX(-50%); background:#333; color:#fff; font-size:0.7rem; padding:0.4rem 0.6rem; border-radius:4px; white-space:nowrap; min-width:180px; line-height:1.6; }}
.sec-tip::after {{ content:""; position:absolute; top:100%; left:50%; transform:translateX(-50%); border:5px solid transparent; border-top-color:#333; }}

.lt {{ font-size:0.72rem; font-weight:500; }}
.lt-fast {{ color:var(--srg-green); }}
.lt-medium {{ color:var(--srg-yellow); }}
.lt-slow {{ color:var(--srg-red); }}

.table-wrap {{ padding:0.8rem 2.5rem 1rem; overflow-x:auto; display:none; }}
.table-wrap.active {{ display:block; }}
.cards-wrap.active .cards {{ display:grid; }}
.cards-wrap:not(.active) .cards {{ display:none; }}
table {{ width:100%; border-collapse:collapse; background:var(--srg-white); border-radius:var(--r); overflow:hidden; border:1px solid var(--srg-border); font-size:0.82rem; }}
thead th {{ background:var(--srg-red); color:var(--srg-white); padding:0.6rem 0.7rem; text-align:left; font-size:0.78rem; font-weight:600; white-space:nowrap; cursor:pointer; user-select:none; position:relative; }}
thead th:hover {{ background:var(--srg-red-dark); }}
thead th .sort-arrow {{ margin-left:4px; font-size:0.65rem; opacity:0.6; }}
thead th.sorted .sort-arrow {{ opacity:1; }}
tbody td {{ padding:0.45rem 0.7rem; vertical-align:middle; border-bottom:1px solid var(--srg-border); }}
.row-even {{ background:var(--srg-white); }}
.row-odd {{ background:var(--srg-bg); }}
.table-thumb {{ width:120px; height:75px; object-fit:cover; object-position:top; border-radius:3px; border:1px solid var(--srg-border); }}
.cell-host {{ font-weight:600; white-space:nowrap; }}
.cell-center {{ text-align:center; }}
.cell-url a {{ color:var(--srg-red); text-decoration:none; }}
.cell-url a:hover {{ text-decoration:underline; }}
.err-text {{ color:var(--srg-red); font-size:0.75rem; }}
.no-img {{ color:var(--srg-gray-light); }}

.pagination {{ display:flex; align-items:center; justify-content:center; gap:0.3rem; padding:0.8rem 2.5rem 1.5rem; flex-wrap:wrap; }}
.pagination button {{ padding:0.3rem 0.7rem; border:1px solid var(--srg-border); border-radius:4px; background:var(--srg-white); font-size:0.8rem; cursor:pointer; font-family:inherit; transition:all 0.12s; }}
.pagination button.active {{ background:var(--srg-red); color:var(--srg-white); border-color:var(--srg-red); }}
.pagination button:hover:not(.active):not(:disabled) {{ border-color:var(--srg-red); color:var(--srg-red); }}
.pagination button:disabled {{ opacity:0.4; cursor:default; }}
.pagination .page-info {{ font-size:0.8rem; color:var(--srg-gray-light); margin:0 0.5rem; }}

.footer {{ text-align:center; padding:1.2rem; color:var(--srg-gray-light); font-size:0.75rem; border-top:2px solid var(--srg-red); margin-top:0.5rem; }}
.footer a {{ color:var(--srg-gray-light); }}

@media(max-width:600px) {{
    .header,.stats,.toolbar,.cards,.table-wrap,.pagination {{ padding-left:1rem; padding-right:1rem; }}
    .cards {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>

<div class="header">
    <div style="display:flex;align-items:center;gap:1rem;">
        {'<img src="' + thumb_b64 + '" alt="Thumbnail" style="height:56px;border-radius:6px;flex-shrink:0;" />' if thumb_b64 else ''}
        <div><h1>{_t['main_title']}</h1><div class="sub">{_t['subtitle']}</div></div>
    </div>
    <div class="header-meta">Erstellt: {ts}<br>Hosts: {total}</div>
</div>

<div class="stats">
    <div class="stat stat-total"><span class="n">{total}</span> Total</div>
    <div class="stat stat-ok"><span class="n">{ok}</span> OK</div>
    <div class="stat stat-redir"><span class="n">{redirects}</span> Redirect</div>
    <div class="stat stat-err"><span class="n">{errors}</span> Fehler</div>
    <div class="stat stat-speed"><span class="n">{avg_load:.1f}s</span> ⌀ Ladezeit</div>
</div>

<div class="toolbar">
    <button class="active" onclick="setView('cards')">Karten</button>
    <button onclick="setView('table')">Tabelle</button>
    <input type="text" class="search-box" placeholder="Suche nach Host, Titel, URL, Fehler, CMS..." oninput="onSearch(this.value)">
    <select class="filter-select" onchange="onFilter(this.value)">
        <option value="all">Alle Status</option>
        <option value="2xx">2xx OK</option>
        <option value="3xx">3xx Redirect</option>
        <option value="4xx">4xx Client Error</option>
        <option value="5xx">5xx Server Error</option>
        <option value="err">Scan-Fehler</option>
        <option value="new">Delta: Neu</option>
        <option value="changed">Delta: Geändert</option>
    </select>
    <select class="filter-select" onchange="onCmsFilter(this.value)" id="cms-filter">
        <option value="all">Alle CMS</option>
    </select>
</div>

<div id="host-filter-bar" style="display:none" class="filter-active-bar">
    <span id="host-filter-chips"></span>
    <span id="host-filter-count" style="font-size:0.8rem;opacity:0.8;"></span>
    <button onclick="clearHostFilter()">✕ Alle aufheben</button>
</div>

<div class="cards-wrap active" id="cards-wrap"><div class="cards" id="cards-container"></div></div>
<div class="table-wrap" id="table-wrap">
    <table>
        <thead><tr id="table-head"></tr></thead>
        <tbody id="table-body"></tbody>
    </table>
</div>
<div class="pagination" id="pagination"></div>

<div class="footer">
    Browser-basierter Check (Playwright/Chromium) · Web Inventory Reporter · {ts}<br>
    {_a['name']} · <a href="mailto:{_a['email']}">{_a['email']}</a>
</div>

<script>
const ALL_DATA = {data_json};
const PER_PAGE = 50;
let view = 'cards';
let filtered = [...ALL_DATA];
let sortCol = null, sortAsc = true;
let page = 1;
let searchTerm = '', statusFilter = 'all', cmsFilter = 'all';
let hostFilters = {{}}; // map type → value, e.g. {{'ip':'1.2.3.4','cdn':'Cloudflare'}}

// Populate CMS filter
(function() {{
  const cms = [...new Set(ALL_DATA.map(r=>r.cms).filter(Boolean))].sort();
  const sel = document.getElementById('cms-filter');
  cms.forEach(c => {{ const o=document.createElement('option'); o.value=c; o.textContent=c; sel.appendChild(o); }});
}})();

function badge(st) {{
  if (!st) return '<span class="badge badge-unknown">N/A</span>';
  const s=parseInt(st)||0;
  const c=s>=200&&s<300?'ok':s>=300&&s<400?'redirect':s>=400&&s<500?'client-err':'server-err';
  return `<span class="badge badge-${{c}}">${{st}}</span>`;
}}
function ltHtml(lt) {{
  if (lt==null) return '—';
  const t=parseFloat(lt), c=t<1.5?'fast':t<3?'medium':'slow';
  return `<span class="lt lt-${{c}}">${{t.toFixed(2)}}s</span>`;
}}
function secHtml(score, detail) {{
  const c=score>=6?'high':score>=4?'mid':'low';
  const keys=Object.keys(detail||{{}});
  const tip=keys.map(k=>`${{detail[k]?'✓':'✗'}} ${{k}}`).join('<br>');
  return `<span class="sec-score sec-${{c}} sec-tooltip">${{score}}/7<span class="sec-tip">${{tip}}</span></span>`;
}}
function deltaHtml(d) {{
  if (!d||d==='unchanged') return '';
  return `<span class="badge badge-${{d}}">${{d==='new'?'Neu':'Geändert'}}</span>`;
}}
function redirHtml(count, chain) {{
  if (!count) return '';
  const tip=(chain||[]).join('&#10;');
  return `<span class="redir-chip" title="${{tip}}">↪ ${{count}}</span>`;
}}
function esc(s) {{ const d=document.createElement('div'); d.textContent=s||''; return d.innerHTML; }}

function applyFilters() {{
  let d = ALL_DATA;
  if (statusFilter !== 'all') {{
    d = d.filter(r => {{
      if (statusFilter==='err')     return !!r.error;
      if (statusFilter==='new')     return r.delta==='new';
      if (statusFilter==='changed') return r.delta==='changed';
      return r.status.startsWith(statusFilter[0]);
    }});
  }}
  if (cmsFilter !== 'all') d = d.filter(r => r.cms === cmsFilter);
  // Apply each active host filter (AND-logic across types)
  Object.entries(hostFilters).forEach(([type, val]) => {{
    d = d.filter(r => {{
      if (type==='ip')  return r.ip === val;
      if (type==='ns')  return (r.nameservers||'').split(',').map(s=>s.trim()).includes(val);
      if (type==='cdn') return r.cdn === val;
      if (type==='asn') return r.asn === val;
      return true;
    }});
  }});
  if (searchTerm) {{
    const q=searchTerm.toLowerCase();
    d = d.filter(r => (r.host+' '+r.title+' '+r.final_url+' '+r.error+' '+r.cms+' '+r.cdn+' '+r.asn+' '+(r.asn_name||'')+' '+(r.scan_state||'')).toLowerCase().includes(q));
  }}
  if (sortCol !== null) {{
    d = [...d].sort((a,b) => {{
      let va=a[sortCol], vb=b[sortCol];
      if (sortCol==='load_time')  {{ va=va??999; vb=vb??999; return sortAsc?va-vb:vb-va; }}
      if (sortCol==='status')     {{ va=parseInt(va)||999; vb=parseInt(vb)||999; return sortAsc?va-vb:vb-va; }}
      if (sortCol==='sec_score')  {{ va=va??0; vb=vb??0; return sortAsc?va-vb:vb-va; }}
      va=(va||'').toString().toLowerCase(); vb=(vb||'').toString().toLowerCase();
      return sortAsc?va.localeCompare(vb):vb.localeCompare(va);
    }});
  }}
  filtered=d; page=1;
  // ── Update filter bar ──
  const bar=document.getElementById('host-filter-bar');
  const _af=Object.entries(hostFilters);
  if (_af.length > 0) {{
    const _lbl={{'ip':'\uD83D\uDDA5 IP','ns':'\uD83C\uDF10 NS','cdn':'CDN','asn':'\uD83C\uDFE2 ASN'}};
    document.getElementById('host-filter-chips').innerHTML = _af.map(([t,v]) =>
      `<span class="f-tag">${{_lbl[t]||t}}: ${{v}}\u2009<span class="f-x" data-clear="${{t}}" title="Filter entfernen">\u00D7</span></span>`
    ).join('\u2002');
    document.getElementById('host-filter-count').textContent = '\u2014 '+d.length+' Treffer';
    bar.style.display='flex';
  }} else {{
    bar.style.display='none';
  }}
  render();
}}

function setHostFilter(type, value) {{
  hostFilters[type] = value;
  applyFilters();
}}
function clearOneFilter(type) {{
  delete hostFilters[type];
  applyFilters();
}}
function clearHostFilter() {{
  hostFilters = {{}};
  applyFilters();
}}

function onSearch(v)    {{ searchTerm=v;    applyFilters(); }}
function onFilter(v)    {{ statusFilter=v;  applyFilters(); }}
function onCmsFilter(v) {{ cmsFilter=v;     applyFilters(); }}
function setView(v) {{
  view=v;
  document.getElementById('cards-wrap').classList.toggle('active', v==='cards');
  document.getElementById('table-wrap').classList.toggle('active', v==='table');
  document.querySelectorAll('.toolbar button').forEach((b,i)=>b.classList.toggle('active',(i===0&&v==='cards')||(i===1&&v==='table')));
  render();
}}
function setPage(p) {{ page=p; render(); window.scrollTo({{top:document.querySelector('.toolbar').offsetTop-10,behavior:'smooth'}}); }}

function render() {{
  const total=filtered.length;
  const pages=Math.max(1,Math.ceil(total/PER_PAGE));
  if (page>pages) page=pages;
  const start=(page-1)*PER_PAGE, end=Math.min(start+PER_PAGE,total);
  const slice=filtered.slice(start,end);

  // ── Cards ──
  document.getElementById('cards-container').innerHTML = slice.map(r => {{
    const img   = r.thumb ? `<img src="${{r.thumb}}" alt="${{esc(r.host)}}" loading="lazy">` : '<div class="no-preview">Kein Preview</div>';
    const ssl   = r.ssl_expiry ? `<span class="card-ssl">🔒 ${{esc(r.ssl_expiry)}}</span>` : '';
    const cms   = r.cms ? `<span class="badge badge-cms">${{esc(r.cms)}}</span>` : '';
    const cdn   = r.cdn ? `<span class="badge badge-cdn filter-link" data-ftype="cdn" data-fval="${{esc(r.cdn)}}" title="Filter: alle Hosts hinter ${{esc(r.cdn)}}">${{esc(r.cdn)}}</span>` : '';
    const sec   = secHtml(r.sec_score, r.sec_detail);
    const redir = redirHtml(r.redirect_count, r.redirect_chain);
    const delta = deltaHtml(r.delta);
    const err    = r.error ? `<div class="card-error">${{esc(r.error.substring(0,120))}}</div>` : '';
    const urlPfx = r.redirect_count > 0 ? '→ ' : '';
    const ipEl  = r.ip ? `<span class="card-ip filter-link" data-ftype="ip" data-fval="${{esc(r.ip)}}" title="Filter: alle Hosts auf IP ${{esc(r.ip)}}&#10;${{esc(r.reverse_dns||'')}}">🖥 ${{esc(r.ip)}}</span>` : '';
    const nsEls = (r.nameservers||'').split(',').filter(s=>s.trim()).map(ns=>`<span class="filter-link" data-ftype="ns" data-fval="${{ns.trim()}}" title="Filter: alle Hosts mit NS ${{ns.trim()}}">${{esc(ns.trim())}}</span>`).join(' · ');
    const nsEl  = nsEls ? (r.ip?'<br>':'')+'<span class="card-ns">🌐 '+nsEls+'</span>' : '';
    const asnEl = r.asn ? `<br><span class="card-asn filter-link" data-ftype="asn" data-fval="${{esc(r.asn)}}" title="Filter: alle Hosts bei ${{esc(r.asn)}}">🏢 ${{esc(r.asn)}}${{r.asn_name ? ' · ' + esc(r.asn_name) : ''}}${{r.hosting_country ? ' (' + esc(r.hosting_country) + ')' : ''}}</span>` : '';
    const dnsBlock = (r.ip || r.nameservers || r.asn) ? `<div class="card-dns">${{ipEl}}${{nsEl}}${{asnEl}}</div>` : '';
    return `<div class="card"><div class="card-image">${{img}}</div><div class="card-body">
      <h3 class="card-title">${{esc(r.host)}} ${{delta}}</h3>
      <div class="card-meta">${{badge(r.status)}} ${{ltHtml(r.load_time)}} ${{ssl}} ${{redir}}</div>
      <div class="card-meta">${{cms}} ${{cdn}} ${{sec}}</div>
      <div class="card-pagetitle">${{esc((r.title||'—').substring(0,60))}}</div>
      <a class="card-url" href="${{esc(r.final_url)}}" target="_blank" rel="noopener">${{urlPfx}}${{esc((r.final_url||'—').substring(0,70))}}</a>
      ${{dnsBlock}}
      ${{err}}</div></div>`;
  }}).join('');

  // ── Table ──
  const cols = [
    ['thumb','Preview',false],['host','Host',true],['status','Status',true],
    ['scan_state','Scan State',true],
    ['load_time','Ladezeit',true],['ssl_expiry','SSL bis',true],['cms','CMS',true],
    ['cdn','CDN',true],
    ['sec_score','Sec.',true],['redirect_count','Redir.',true],
    ['ip','IP',true],['asn','ASN',true],['nameservers','Nameserver',true],
    ['title','Titel',true],['final_url','URL',true],['delta','Delta',true],['error','Fehler',true]
  ];
  document.getElementById('table-head').innerHTML = cols.map(([key,label,sortable]) => {{
    const arrow=sortCol===key?(sortAsc?'▲':'▼'):'⇅';
    const cls=sortCol===key?' class="sorted"':'';
    return sortable ? `<th${{cls}} onclick="doSort('${{key}}')">${{label}}<span class="sort-arrow">${{arrow}}</span></th>` : `<th>${{label}}</th>`;
  }}).join('');

  document.getElementById('table-body').innerHTML = slice.map((r,i) => {{
    const cls=i%2===0?'row-even':'row-odd';
    const img=r.thumb?`<img src="${{r.thumb}}" class="table-thumb">`:'<span class="no-img">—</span>';
    const err=r.error?`<span class="err-text">${{esc(r.error.substring(0,80))}}</span>`:'—';
    const chain=(r.redirect_chain||[]).join('\\n');
    return `<tr class="${{cls}}">
      <td>${{img}}</td>
      <td class="cell-host">${{esc(r.host)}}</td>
      <td class="cell-center">${{badge(r.status)}}</td>
      <td class="cell-center"><code style="font-size:0.72rem">${{esc(r.scan_state||'—')}}</code></td>
      <td class="cell-center">${{ltHtml(r.load_time)}}</td>
      <td class="cell-center">${{esc(r.ssl_expiry||'—')}}</td>
      <td class="cell-center">${{r.cms?`<span class="badge badge-cms">${{esc(r.cms)}}</span>`:'—'}}</td>
      <td class="cell-center">${{r.cdn?`<span class="badge badge-cdn filter-link" data-ftype="cdn" data-fval="${{esc(r.cdn)}}" title="Filter: ${{esc(r.cdn)}}">${{esc(r.cdn)}}</span>`:'—'}}</td>
      <td class="cell-center">${{secHtml(r.sec_score,r.sec_detail)}}</td>
      <td class="cell-center">${{r.redirect_count?`<span class="redir-chip" title="${{esc(chain)}}">${{r.redirect_count}}x</span>`:'—'}}</td>
      <td class="cell-ip"><span class="${{r.ip?'filter-link':''}}" data-ftype="ip" data-fval="${{esc(r.ip||'')}}" title="${{esc(r.reverse_dns||'')}}${{r.ip?' → Filter: '+esc(r.ip):''}}">${{esc(r.ip||'—')}}</span></td>
      <td class="cell-asn"><span class="${{r.asn?'filter-link':''}}" data-ftype="asn" data-fval="${{esc(r.asn||'')}}" title="${{esc(r.asn_name||'')}}${{r.asn?' → Filter: '+esc(r.asn):''}}">${{esc(r.asn||'—')}}</span></td>
      <td class="cell-ns">${{esc((r.nameservers||'—').substring(0,50))}}</td>
      <td>${{esc((r.title||'—').substring(0,50))}}</td>
      <td class="cell-url"><a href="${{esc(r.final_url)}}" target="_blank">${{esc((r.final_url||'—').substring(0,55))}}</a></td>
      <td class="cell-center">${{deltaHtml(r.delta)}}</td>
      <td>${{err}}</td></tr>`;
  }}).join('');

  // ── Pagination ──
  const pg=document.getElementById('pagination');
  if (pages<=1) {{ pg.innerHTML=`<span class="page-info">${{total}} Hosts</span>`; return; }}
  let h=`<button ${{page===1?'disabled':''}} onclick="setPage(${{page-1}})">‹</button>`;
  const maxBtns=9;
  let pStart=Math.max(1,page-Math.floor(maxBtns/2));
  let pEnd=Math.min(pages,pStart+maxBtns-1);
  if (pEnd-pStart<maxBtns-1) pStart=Math.max(1,pEnd-maxBtns+1);
  if (pStart>1) h+=`<button onclick="setPage(1)">1</button><span class="page-info">…</span>`;
  for (let p=pStart;p<=pEnd;p++) h+=`<button class="${{p===page?'active':''}}" onclick="setPage(${{p}})">${{p}}</button>`;
  if (pEnd<pages) h+=`<span class="page-info">…</span><button onclick="setPage(${{pages}})">${{pages}}</button>`;
  h+=`<button ${{page===pages?'disabled':''}} onclick="setPage(${{page+1}})">›</button>`;
  h+=`<span class="page-info">${{start+1}}–${{end}} von ${{total}}</span>`;
  pg.innerHTML=h;
}}

function doSort(col) {{
  if (sortCol===col) sortAsc=!sortAsc; else {{sortCol=col; sortAsc=true;}}
  applyFilters();
}}

// ── Click handlers for filter links and × chips ──
document.addEventListener('click', function(e) {{
  // Remove single filter via × chip
  const cx = e.target.closest('.f-x[data-clear]');
  if (cx) {{ e.stopPropagation(); clearOneFilter(cx.dataset.clear); return; }}
  // Set / toggle a host filter
  const el = e.target.closest('.filter-link[data-ftype]');
  if (!el) return;
  const type = el.dataset.ftype;
  const val  = el.dataset.fval;
  if (!val || val === '—') return;
  e.stopPropagation();
  // Toggle: clicking the active value of the same type removes just that type
  if (hostFilters[type] === val) {{
    clearOneFilter(type);
  }} else {{
    setHostFilter(type, val);
  }}
}});

render();
</script>
</body></html>"""

    with open(OUTPUT_DIR / "report.html", "w", encoding="utf-8") as f:
        f.write(html)
    log("HTML report written.")


# ── JSON Report (AI-readable) ────────────────────────────────────────────────

def write_json_report(data: list):
    ts = datetime.now().isoformat()
    total = len(data)
    ok       = sum(1 for d in data if str(d.get("status", "")).startswith("2"))
    redir    = sum(1 for d in data if str(d.get("status", "")).startswith("3"))
    cli_err  = sum(1 for d in data if str(d.get("status", "")).startswith("4"))
    srv_err  = sum(1 for d in data if str(d.get("status", "")).startswith("5"))
    errors   = sum(1 for d in data if d.get("error"))
    load_times = [float(d["load_time"]) for d in data if d.get("load_time")]

    report = {
        "meta": {
            "generated_at": ts,
            "tool":         "Web Inventory Reporter",
            "scan_method":  "Browser-basierter Check (Playwright/Chromium)",
            "total_hosts":  total,
            "summary": {
                "status_2xx":   ok,
                "status_3xx":   redir,
                "status_4xx":   cli_err,
                "status_5xx":   srv_err,
                "scan_errors":  errors,
                "avg_load_s":   round(sum(load_times) / len(load_times), 2) if load_times else None,
                "min_load_s":   round(min(load_times), 2) if load_times else None,
                "max_load_s":   round(max(load_times), 2) if load_times else None,
            },
        },
        "hosts": [],
    }

    for d in data:
        sh = d.get("security_headers", {})
        report["hosts"].append({
            "host":            d["host"],
            "ip":              d.get("ip", ""),
            "reverse_dns":     d.get("reverse_dns", ""),
            "nameservers":     [ns.strip() for ns in d.get("nameservers", "").split(",") if ns.strip()],
            "asn":             d.get("asn", ""),
            "asn_name":        d.get("asn_name", ""),
            "hosting_country": d.get("hosting_country", ""),
            "scan_state":      d.get("scan_state", ""),
            "scan_time":       d.get("scan_time", ""),
            "status":          d.get("status", ""),
            "final_url":       d.get("final_url", ""),
            "title":           d.get("title", ""),
            "load_time_s":     d.get("load_time") or None,
            "ssl_expiry":      d.get("ssl_expiry", ""),
            "cert_subject_cn": d.get("cert_subject_cn", ""),
            "cert_san":        [s.strip() for s in d.get("cert_san", "").split(",") if s.strip()],
            "cms":             d.get("cms", ""),
            "cdn":             d.get("cdn", ""),
            "registrar":       d.get("registrar", ""),
            "domain_expiry":   d.get("domain_expiry", ""),
            "delta":           d.get("delta", ""),
            "error":           d.get("error", ""),
            "redirects": {
                "count": len(d.get("redirect_chain", [])),
                "chain": [
                    {"url": r["url"], "status": r["status"]}
                    for r in d.get("redirect_chain", [])
                ],
            },
            "security_headers": {
                "score":   sh.get("_score", 0),
                "max":     len(SEC_HEADERS),
                "present": [h for h in SEC_HEADERS if sh.get(h)],
                "missing": [h for h in SEC_HEADERS if not sh.get(h)],
            },
        })

    with open(OUTPUT_DIR / "report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, default=str)
    log("JSON report written.")


# ── Parallel worker ───────────────────────────────────────────────────────────

_print_lock = threading.Lock()


def _worker(host_queue: queue.Queue, results: list, counter: list, total: int):
    """Each worker owns its own sync_playwright + browser + page."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx     = browser.new_context(viewport={"width": 1280, "height": 800})
        page    = ctx.new_page()
        page.set_default_timeout(PAGE_TIMEOUT)
        try:
            while True:
                try:
                    idx, host = host_queue.get_nowait()
                except queue.Empty:
                    break
                result = scan_host(page, browser, host)
                results[idx] = result
                with _print_lock:
                    counter[0] += 1
                    n      = counter[0]
                    status = result["status"] or "ERR"
                    lt     = f'{result["load_time"]}s' if result["load_time"] else ""
                    cms    = f' [{result["cms"]}]' if result.get("cms") else ""
                    sec    = f' sec:{result.get("security_headers", {}).get("_score", 0)}/7'
                    redir  = f' ↪{len(result["redirect_chain"])}' if result.get("redirect_chain") else ""
                    print(f"  [{n}/{total}] {host} → {status} {lt}{redir}{cms}{sec}")
                host_queue.task_done()
        finally:
            ctx.close()
            browser.close()


# ── Main ─────────────────────────────────────────────────────────────────────

def main(input_file):
    hosts    = load_hosts(input_file)
    previous = load_previous_results()
    total    = len(hosts)
    workers  = min(WORKERS, total) if total else 1

    print(f"Web Inventory Reporter — {total} Hosts | {workers} Workers")
    print("─" * 50)

    # Fill queue
    host_queue = queue.Queue()
    for i, host in enumerate(hosts):
        host_queue.put((i, host))

    results = [None] * total
    counter = [0]

    threads = [
        threading.Thread(target=_worker, args=(host_queue, results, counter, total), daemon=True)
        for _ in range(workers)
    ]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    results = [r for r in results if r is not None]
    if not results:
        print("Keine Hosts gefunden.")
        return

    # Delta
    for r in results:
        r["delta"] = compute_delta(r, previous)

    # Mark hosts that disappeared since last scan
    current_hosts = {r["host"] for r in results}
    gone = [h for h in previous if h not in current_hosts]
    if gone:
        print(f"  Weggefallen seit letztem Scan: {', '.join(gone)}")

    save_results_json(results)

    # Sanitize surrogate characters that some pages produce (broken encodings).
    # Python's UTF-8 codec refuses to write them — replace with U+FFFD (replacement char).
    # Recursive so nested structures (security_headers dict, redirect_chain list) are covered too.
    def _clean(v):
        if isinstance(v, str):
            return v.encode("utf-8", errors="replace").decode("utf-8")
        if isinstance(v, dict):
            return {k: _clean(val) for k, val in v.items()}
        if isinstance(v, list):
            return [_clean(item) for item in v]
        return v
    for r in results:
        for k in list(r.keys()):
            r[k] = _clean(r[k])

    print(f"\n{'─'*50}")
    print(f"Reports für {len(results)} Hosts ...")
    write_csv(results)
    write_excel(results)
    write_html(results)
    write_json_report(results)
    print("  ✓ output/report.csv")
    print("  ✓ output/report.xlsx")
    print("  ✓ output/report.html")
    print("  ✓ output/report.json")
    print("─" * 50)
    print("Fertig!")


if __name__ == "__main__":
    default_file = CFG["general"]["domains_file"]
    input_file = sys.argv[1] if len(sys.argv) >= 2 else default_file
    main(input_file)